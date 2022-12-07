import openpyxl
import requests
import json

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("/Users/aikyab/Documents/DWebApp/backend/college_data.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

def post_universities():
        
        fields = []
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            fields.append(col[1].value.lower())

        for row in range(2,13):
            i = 0
            univ = {}
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                field_val = col[row].value
                if field_val:
                    field_val = str(field_val)
                
                if field_val and fields[i] in ["black_aa","latinx","aapi","white","mixed_race","unknown","international"]:
                    field_val = float(field_val)

                univ[fields[i]] = field_val
                i+=1
            try:
                response = requests.post("http://localhost:8000/university", json=univ, verify=False)
                print(response.status_code ,univ["name"])
            except Exception as exp:
                continue
            
def get_universities():
    response = requests.get("http://localhost:8000/university/")
    data = json.loads(response.text)
    print(data)

def get_university_by_id(id):
    response = requests.get(f"http://localhost:8000/university/getById/{id}")
    print(json.loads(response.text))


if __name__ == '__main__':
    
    post_universities()
    # get_universities()
    # get_university_by_id(248)
    pass
