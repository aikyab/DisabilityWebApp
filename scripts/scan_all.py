import openpyxl
import requests
import json
from collections import Counter, defaultdict
import pickle

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("college_data.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active


def post_universities():
    univ = {}
    fields = []
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        fields.append(col[1].value.lower())

    for row in range(3,65):
        i = 0
        # for col in dataframe1.iter_cols(1, dataframe1.max_column):
        name = "A{}".format(row)
        codes = "AQ{}".format(row)
        if dataframe1[codes].value:
            univ[dataframe1[name].value] = str(dataframe1[codes].value).strip().split()
    
    with open('universities.pkl', 'wb') as f:
        pickle.dump(univ, f)
        

        #     if field_val:
        #         field_val = str(field_val)
            
        #     if field_val and fields[i] in ["black_aa","latinx","aapi","white","mixed_race","unknown","international"]:
        #         field_val = float(field_val)

        #     univ[fields[i]] = field_val
        # i+=1
        # try:
        #     response = requests.post("http://localhost:8000/university", json=univ, verify=False)
        #     print(response.status_code ,univ["name"])
        # except Exception:
        #     continue

def scan_all_codes():

    def scan_rows(up,down,field_list,fields):
        i = 0
        for row in range(up,down):
            need = "{}{}".format("B", row)
            no_need = "{}{}".format("C", row)
            fields[field_list[i]]["Essential"] = str(codes[need].value)
            fields[field_list[i]]["Don't Need"] = str(codes[no_need].value)
        
            i+=1

    codes = openpyxl.load_workbook("codes.xlsx")
    codes = codes.active

    fields = defaultdict(dict)

    gen_acc = ["asl_interpreters","note_takers","recorded_lectures",
    "test_acc","assist_ldevices","rem_arch_barriers","course_subwaivers","alt_text_materials",
    "preferential_setting","adaptive_equipment","assistive_tech"]

    supports = ["writing","math","career_center","study_skills","tutoring"]

    o_supports = ["peer_mentoring","faculty_mentoring","social_skills","summer_bridge","prog_iddstudents"]

    for row in range(22):
        key = "A{}".format(row)
        val = "B{}".format(row)
        if row in [3,4,5]:
            fields["university_type"][codes[key].value] = str(codes[val].value)
        elif row in [7,8,9,10]:
            fields["geo_setting"][codes[key].value] = str(codes[val].value)
        elif row in [12,13,14,15]:
            fields["school_size"][codes[key].value] = str(codes[val].value)
        elif row in [17,18,19,20,21]:
            fields["selectivity"][codes[key].value.replace('\xa0', '')] = str(codes[val].value).replace('\xa0', '')

    scan_rows(24,35,gen_acc,fields)
    scan_rows(37,42,supports,fields)
    scan_rows(43,48,o_supports,fields)
    
    
    # //save dictionary in python pickle? 
    with open('codes.pkl', 'wb') as f:
        pickle.dump(fields, f)

def get_universities():
    response = requests.get("http://localhost:8000/university/")
    data = json.loads(response.text)
    print(data)

def get_university_by_id(id):
    response = requests.get(f"http://localhost:8000/university/getById/{id}")
    print(json.loads(response.text))


if __name__ == '__main__':

    post_universities()
    # get_universities()
